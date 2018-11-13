from openpyxl import Workbook
from openpyxl import load_workbook
import os
import urllib.request
import threading


class opreate():
    def __init__(self):
        wb = load_workbook("d_card_dis.xlsx")
        ws = wb.get_sheet_by_name('@卡牌表')

        print(wb.sheetnames)

        id = ws['C1']

        i = 1

        while(id.value != '' and id.value != None):
            print(id.value)

            type = ws['D'+str(i)]
            pinzhi = ws['F'+str(i)]
            zhiye = ws['J'+str(i)]

            typeSTR = type.value
            pinzhiSTR = pinzhi.value
            zhiyeSTR = zhiye.value

            if typeSTR == '仆从':
                type.value = '3'
            elif typeSTR == '法术':
                type.value = '1'
            elif typeSTR == '武器':
                type.value = '4'
            elif typeSTR == '奥秘':
                type.value = '3'
            elif typeSTR == '英雄':
                type.value = '3'
            elif typeSTR == '英雄技能':
                type.value = '1'

            if pinzhiSTR == '普通':
                pinzhi.value = '1'
            elif pinzhiSTR == '免费':
                pinzhi.value = '1'
            elif pinzhiSTR == '精良':
                pinzhi.value = '2'
            elif pinzhiSTR == '优秀':
                pinzhi.value = '2'
            elif pinzhiSTR == '史诗':
                pinzhi.value = '3'
            elif pinzhiSTR == '传说':
                pinzhi.value = '4'

            if zhiyeSTR == '法师':
                zhiye.value = '0'
            elif zhiyeSTR == '猎人':
                zhiye.value = '1'
            elif zhiyeSTR == '圣骑士':
                zhiye.value = '2'
            elif zhiyeSTR == '战士':
                zhiye.value = '3'
            elif zhiyeSTR == '德鲁伊':
                zhiye.value = '4'
            elif zhiyeSTR == '术士':
                zhiye.value = '5'
            elif zhiyeSTR == '萨满':
                zhiye.value = '6'
            elif zhiyeSTR == '牧师':
                zhiye.value = '7'
            elif zhiyeSTR == '潜行者':
                zhiye.value = '8'
            elif zhiyeSTR == '中立':
                zhiye.value = '10'


            i+= 1
            id = ws['C'+str(i)]

        wb.save('d_card_dis.xlsx')

oo = opreate()